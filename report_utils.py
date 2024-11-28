import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import logging

logger = logging.getLogger(__name__)

# Explanations for each category and subcategory
EXPLANATIONS = {
    "Validation": "This confirms whether the email format and domain are valid.",
    "WHOIS": {
        "Registrar": "The organization responsible for managing the domain registration.",
        "Creation Date": "The date when the domain was first registered.",
        "Expiration Date": "The date when the domain registration is set to expire.",
        "Organization": "The entity that owns or manages the domain.",
    },
    "Reverse DNS": {
        "Hostname": "The name assigned to an IP address, used for reverse lookups.",
        "Aliases": "Alternative names associated with the IP address.",
        "IPs": "IP addresses linked to the hostname or domain.",
    },
    "Mail Server Details": {
        "MX Records": "Mail Exchange servers responsible for receiving email for the domain.",
        "SPF": "A record that prevents email spoofing by verifying sender servers.",
        "DMARC": "A policy that enforces email authentication using SPF and DKIM.",
    },
    "VirusTotal Reputation": "Results from VirusTotal's analysis of the domain's reputation.",
    "HIBP Results": "Check if the email has been involved in any data breaches.",
    "Social Media Accounts": {
        "Found Accounts": "Social media profiles associated with the email or username.",
        "Failed Checks": "Platforms where social media lookup could not be completed.",
    },
    "AI Summary": "A comprehensive AI-generated summary of the analysis results.",
    "Detailed Report": "A detailed AI-written descriptive report of the analysis results.",
}


def apply_formatting(sheet):
    """Apply formatting to the Excel sheet, including Columns A, B, and C."""
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            # Bold headers and section titles
            if cell.row == 1:  # First row is the header
                cell.font = Font(bold=True)
            # Align text and wrap text in Columns B and C
            if cell.column in [2, 3]:  # Column B (Details) and Column C (Description)
                cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
            else:  # General alignment for other columns (e.g., Column A)
                cell.alignment = Alignment(horizontal="left", vertical="top")

    # Adjust column widths for readability
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 110
    sheet.column_dimensions["C"].width = 80


def save_results_to_excel(email, details, ai_summary, detailed_report):
    """Save all OSINT results to an Excel file, including the detailed AI-written report."""
    logger.info("Starting to save results to Excel...")

    username = email.split('@')[0].replace('.', '')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OSINT Results"

    # Add headers
    sheet.append(["Category", "Details", "Description"])
    sheet.append([])  # Blank row for separation

    # Add Detailed Report
    detailed_intro = (
        "This section contains a detailed AI-generated report that provides "
        "a comprehensive description of the findings based on the analysis."
    )
    sheet.append(["Detailed Report", detailed_report, detailed_intro])
    sheet.append([])  # Add an empty row to separate sections

    # Populate the rest of the results
    added_ai_summary = False  # Flag to track if AI Summary is already added
    for key, value in details.items():
        if key == "VirusTotal Reputation":
            # Handle VirusTotal Reputation Section
            vt_data = value
            if isinstance(vt_data, dict):
                for subkey, subvalue in vt_data.items():
                    explanation = EXPLANATIONS.get("VirusTotal Reputation", "Details from VirusTotal reputation analysis.")
                    sheet.append([f"VirusTotal Reputation - {subkey}", str(subvalue), explanation])
            else:
                explanation = EXPLANATIONS.get("VirusTotal Reputation", "Details from VirusTotal reputation analysis.")
                sheet.append(["VirusTotal Reputation", str(vt_data), explanation])
        elif key == "HIBP Results":
            # Handle HIBP Results Section
            hibp_data = value
            if isinstance(hibp_data, list):  # If breaches are found, list them
                explanation = EXPLANATIONS.get("HIBP Results", "Details of email breaches from HIBP analysis.")
                sheet.append([f"HIBP Results - Breaches", ", ".join(hibp_data), explanation])
            else:
                explanation = EXPLANATIONS.get("HIBP Results", "Check if the email has been involved in any data breaches.")
                sheet.append(["HIBP Results", str(hibp_data), explanation])
        elif isinstance(value, dict):
            for subkey, subvalue in value.items():
                explanation = EXPLANATIONS.get(key, {}).get(subkey, f"Explanation for {key} - {subkey}")
                sheet.append([f"{key} - {subkey}", str(subvalue), explanation])
        else:
            if key == "AI Summary" and added_ai_summary:
                continue  # Avoid duplicate AI Summary
            explanation = EXPLANATIONS.get(key, f"Explanation for {key}")
            sheet.append([key, str(value), explanation])
            if key == "AI Summary":
                added_ai_summary = True

    # Add AI summary explicitly if it wasn't added
    if not added_ai_summary:
        ai_explanation = EXPLANATIONS.get("AI Summary", "Comprehensive summary of the results.")
        sheet.append(["AI Summary", ai_summary, ai_explanation])

    # Apply formatting
    apply_formatting(sheet)

    # Handle locked files
    base_filename = f"osint_results_{username}.xlsx"
    file_name = base_filename
    file_appendix = 1

    while True:
        try:
            workbook.save(file_name)
            logger.info(f"Results saved to {file_name}")
            return file_name
        except PermissionError:
            logger.warning(f"File '{file_name}' is locked. Attempting to save with a new name...")
            file_name = f"osint_results_{username}_{file_appendix}.xlsx"
            file_appendix += 1
